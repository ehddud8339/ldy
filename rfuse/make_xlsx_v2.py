import os
import re
import sys
from pathlib import Path
from collections import defaultdict, Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side

FILENAME_RE = re.compile(
    r'^(?P<cpus>[\d,]+)_(?P<workload>[A-Za-z]+)_(?P<bs>\d+k)_(?P<numjobs>\d+)\.log$'
)

def _num_with_suffix_to_float(s: str) -> float:
    s = s.strip()
    m = re.match(r'([0-9]*\.?[0-9]+)\s*([kKmMgG]?)', s)
    if not m:
        return float("nan")
    val, suf = float(m.group(1)), m.group(2).lower()
    return val * {"k": 1e3, "m": 1e6, "g": 1e9}.get(suf, 1)

def parse_fio_file(path: Path) -> dict:
    text = path.read_text(errors="ignore")

    # IOPS
    iops = None
    m = re.search(r"IOPS\s*=\s*([0-9\.]+[kKmMgG]?)", text)
    if m: iops = _num_with_suffix_to_float(m.group(1))

    # BW(MB/s)
    bw = None
    m = re.search(r"\(([0-9\.]+)\s*MB/s\)", text)
    if m: bw = float(m.group(1))
    else:
        m = re.search(r"BW\s*=\s*([0-9\.]+)\s*MiB/s", text, re.I)
        if m: bw = float(m.group(1)) * 1.048576

    # lat_avg (prefer lat, fallback clat); normalize to usec
    lat_avg = None
    m = re.search(r"lat\s*\((nsec|usec|msec)\).*?avg\s*=\s*([0-9\.]+)", text, re.S | re.I) \
        or re.search(r"clat\s*\((nsec|usec|msec)\).*?avg\s*=\s*([0-9\.]+)", text, re.S | re.I)
    if m:
        unit, val = m.group(1).lower(), float(m.group(2))
        if unit == "nsec": lat_avg = val / 1000.0
        elif unit == "usec": lat_avg = val
        elif unit == "msec": lat_avg = val * 1000.0

    return {"IOPS": iops, "BW(MB/s)": bw, "lat_avg": lat_avg}

def bs_key(bs: str):
    m = re.match(r"(\d+)\s*k$", bs.strip(), re.I)
    return int(m.group(1)) if m else 10**9

def workload_key(w):
    order = ["randread", "randwrite", "read", "write"]
    return order.index(w) if w in order else len(order)

def select_target_group(cpus_list, start_char, preferred):
    """
    Pick a 4-length cpus string starting with start_char.
    Priority:
      1) If any of 'preferred' exist (in order), return the first found.
      2) Else any cpus that startswith start_char and has at least 4 entries:
         - choose one with exactly 4 entries if available (tie-breaker: smallest numerically)
         - else choose the one with the largest available length, then smallest numerically.
    """
    candidates = [s for s in cpus_list if s.startswith(start_char)]
    # 1) preferred
    for p in preferred:
        if p in candidates:
            return p
    # 2) general
    four = [s for s in candidates if len(s.split(",")) == 4]
    if four:
        return sorted(four, key=lambda s: [int(x) for x in s.split(",")])[0]
    longer = [s for s in candidates if len(s.split(",")) > 4]
    if longer:
        # pick shortest >4, then numeric
        shortest_len = min(len(s.split(",")) for s in longer)
        shortest = [s for s in longer if len(s.split(",")) == shortest_len]
        return sorted(shortest, key=lambda s: [int(x) for x in s.split(",")])[0]
    # as last resort, try any >=4 length starting with start_char (shouldn't reach here if none exist)
    return None

def prefixes(cg: str, upto: int):
    parts = cg.split(",")
    # Ensure we can make exactly upto prefixes; if not enough parts, pad with last
    if len(parts) < upto:
        parts = parts + [parts[-1]] * (upto - len(parts))
    return [",".join(parts[:k]) for k in range(1, upto+1)]

def main(log_dir=".", output_path=None):
    base = Path(log_dir).resolve()
    logs = [p for p in base.glob("*.log") if FILENAME_RE.match(p.name)]
    if not logs:
        print(f"[WARN] No .log files found in {base}")
        return

    parsed = []
    for p in logs:
        m = FILENAME_RE.match(p.name)
        d = parse_fio_file(p)
        d.update(m.groupdict())
        d["numjobs"] = int(d["numjobs"])
        parsed.append(d)

    df = pd.DataFrame(parsed)

    # Available cpus sets
    cpus_list = sorted(df["cpus"].unique(), key=lambda s: [int(x) for x in s.split(",")])

    # Choose BUSY & IDLE target group
    busy_group = select_target_group(cpus_list, "0", preferred=["0,2,4,6", "0,2,4,8"])
    idle_group = select_target_group(cpus_list, "8", preferred=["8,10,12,14"])

    if not busy_group and not idle_group:
        print("[ERROR] No BUSY/IDLE 4-length cpu groups found in logs.")
        return

    workloads = sorted(df["workload"].unique(), key=workload_key)
    bs_list = sorted(df["bs"].unique(), key=bs_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    bold = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 1
    written_tables = 0

    for cpus in [g for g in [busy_group, idle_group] if g]:
        subs = prefixes(cpus, 4)
        for workload in workloads:
            for bs in bs_list:
                # Header
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
                ws.cell(current_row, 2, f"{cpus}_{workload}_{bs}").font = bold
                ws.cell(current_row, 2).alignment = center
                current_row += 1

                # numjobs header
                for i, nj in enumerate([1,2,3,4], start=2):
                    c = ws.cell(current_row, i, nj)
                    c.font = bold; c.alignment = center; c.border = border
                current_row += 1

                # rows
                for r_name in ["IOPS", "BW(MB/s)", "lat_avg"]:
                    ws.cell(current_row, 1, r_name).font = bold
                    ws.cell(current_row, 1).alignment = right
                    ws.cell(current_row, 1).border = border

                    for i, nj in enumerate([1,2,3,4], start=2):
                        subset = subs[nj - 1]
                        subdf = df[
                            (df["cpus"] == subset) &
                            (df["workload"] == workload) &
                            (df["bs"] == bs) &
                            (df["numjobs"] == nj)
                        ]
                        val = subdf[r_name].iloc[0] if not subdf.empty else None
                        c = ws.cell(current_row, i, val)
                        c.alignment = center; c.border = border
                    current_row += 1

                # spacing
                current_row += 3
                written_tables += 1

    ws.column_dimensions["A"].width = 14
    for col in ["B","C","D","E"]:
        ws.column_dimensions[col].width = 12

    out = Path(output_path) if output_path else base / f"{base.name}_summary.xlsx"
    wb.save(out)
    print(f"[OK] Created {written_tables} tables -> {out}")
    if busy_group:
        print(f"[INFO] BUSY group used: {busy_group}")
    if idle_group:
        print(f"[INFO] IDLE group used: {idle_group}")

if __name__ == "__main__":
    log_dir = sys.argv[1] if len(sys.argv) > 1 else "."
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    main(log_dir, output_path)


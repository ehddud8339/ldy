import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# === 실험 환경 설정 ===
BASE_LOG_DIR = "log_files/rfuse"

WORKLOADS     = ["read", "write", "randread", "randwrite"]
CPU_SETS      = ["0-39", "0-19", "0-9", "0-4", "0"]         # 행 순서
NUMJOBS_LIST  = ["1", "2", "4", "8", "16", "32"]            # 열 순서

# 여러 개의 bs 필터를 허용 (소문자 기준). None이면 전체 수집
BS_FILTERS = ["4k", "128k"]          # 예: ["4k","128k"] / ["128k"] / None

OUTPUT_XLSX = "rfuse_cpu_scaling_summary.xlsx"


# === fio 로그 파싱 ===
def parse_fio_log(filepath, workload):
    # randread/randwrite도 출력 섹션은 read:/write:
    section_key = "read" if workload in ["read", "randread"] else "write"

    iops_val = None
    tail_val_usec = None
    clat_unit = None  # nsec/usec/msec

    with open(filepath, "r", errors="ignore") as f:
        for line in f:
            s = line.strip()

            # IOPS=... (k/M/G 접미사 처리)
            m_iops = re.search(rf"{section_key}:\s+IOPS=([0-9]+(\.[0-9]+)?[kMG]?)", s)
            if m_iops and iops_val is None:
                raw = m_iops.group(1)
                mult = 1.0
                if raw.endswith('k'):
                    mult = 1e3;  raw = raw[:-1]
                elif raw.endswith('M'):
                    mult = 1e6;  raw = raw[:-1]
                elif raw.endswith('G'):
                    mult = 1e9;  raw = raw[:-1]
                try:
                    iops_val = float(raw) * mult
                except ValueError:
                    pass

            # clat 단위 탐지
            m_unit = re.search(r"clat percentiles \((nsec|usec|msec)\):", s)
            if m_unit:
                clat_unit = m_unit.group(1)

            # 99.99th
            m_p = re.search(r"99\.99th=\[\s*([0-9]+)\s*\]", s)
            if m_p and tail_val_usec is None and clat_unit is not None:
                raw_tail = float(m_p.group(1))
                if clat_unit == "usec":
                    tail_val_usec = raw_tail
                elif clat_unit == "msec":
                    tail_val_usec = raw_tail * 1000.0
                elif clat_unit == "nsec":
                    tail_val_usec = raw_tail / 1000.0

    return iops_val, tail_val_usec


# === 시트 작성 ===
def write_workload_bs_sheet(wb, workload, bs_token, data_iops, data_tail):
    title = f"{workload}_bs{bs_token}"
    ws = wb.create_sheet(title=title)

    left_label_col    = 1                  # cpus: 라벨
    left_block_start  = 2                  # IOPS 데이터 시작 열
    right_label_col   = left_block_start + len(NUMJOBS_LIST) + 1
    right_block_start = right_label_col + 1

    bold_center  = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # === 병합 헤더 (키워드 인자 사용) ===
    ws.merge_cells(
        start_row=1,
        start_column=left_block_start,
        end_row=1,
        end_column=left_block_start + len(NUMJOBS_LIST) - 1
    )
    ws.cell(row=1, column=left_block_start).value = f"[{workload}]_IOPS (bs={bs_token})"
    ws.cell(row=1, column=left_block_start).font = bold_center
    ws.cell(row=1, column=left_block_start).alignment = center_align

    ws.merge_cells(
        start_row=1,
        start_column=right_block_start,
        end_row=1,
        end_column=right_block_start + len(NUMJOBS_LIST) - 1
    )
    ws.cell(row=1, column=right_block_start).value = f"[{workload}]_99.99_tail_latency(usec)"
    ws.cell(row=1, column=right_block_start).font = bold_center
    ws.cell(row=1, column=right_block_start).alignment = center_align

    # === 2행: numjobs 헤더 ===
    ws.cell(row=2, column=left_label_col).value = ""   # 라벨 칸 비움
    for idx, nj in enumerate(NUMJOBS_LIST):
        c = left_block_start + idx
        ws.cell(row=2, column=c).value = nj
        ws.cell(row=2, column=c).font = bold_center
        ws.cell(row=2, column=c).alignment = center_align

    ws.cell(row=2, column=right_label_col).value = ""
    for idx, nj in enumerate(NUMJOBS_LIST):
        c = right_block_start + idx
        ws.cell(row=2, column=c).value = nj
        ws.cell(row=2, column=c).font = bold_center
        ws.cell(row=2, column=c).alignment = center_align

    # === 본문 ===
    for r_i, cs in enumerate(CPU_SETS):
        r = 3 + r_i
        # 라벨
        ws.cell(row=r, column=left_label_col).value  = f"cpus:{cs}"
        ws.cell(row=r, column=left_label_col).font   = bold_center
        ws.cell(row=r, column=right_label_col).value = f"cpus:{cs}"
        ws.cell(row=r, column=right_label_col).font  = bold_center

        # 값 채우기
        for idx, nj in enumerate(NUMJOBS_LIST):
            ws.cell(row=r, column=left_block_start  + idx).value = data_iops.get(cs, {}).get(nj)
            ws.cell(row=r, column=right_block_start + idx).value = data_tail.get(cs, {}).get(nj)

    # 열 너비
    max_col = right_block_start + len(NUMJOBS_LIST)
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

# === 메인 ===
def main():
    # data[workload][bs][cpu_set][numjobs] = value
    data_iops = {wl: {} for wl in WORKLOADS}
    data_tail = {wl: {} for wl in WORKLOADS}

    # 허용 bs 집합
    bs_allow = None if BS_FILTERS is None else set([b.lower() for b in BS_FILTERS])

    # 디렉토리 스캔: cpus_0-4, cpus_0, ...
    for d in os.listdir(BASE_LOG_DIR):
        full_dir = os.path.join(BASE_LOG_DIR, d)
        if not os.path.isdir(full_dir):
            continue
        m_dir = re.match(r"cpus_(.+)$", d)
        if not m_dir:
            continue
        cpu_set = m_dir.group(1)
        if cpu_set not in CPU_SETS:
            continue

        for fname in os.listdir(full_dir):
            if not fname.endswith(".log"):
                continue
            # 파일 패턴: randread_bs128k_njs32.log
            m = re.match(r"(read|write|randread|randwrite)_bs([0-9]+[KkMmGg])_njs([0-9]+)\.log$", fname)
            if not m:
                continue
            wl  = m.group(1)
            bs  = m.group(2).lower()   # '128k'
            njs = m.group(3)

            if wl not in WORKLOADS or njs not in NUMJOBS_LIST:
                continue
            if bs_allow is not None and bs not in bs_allow:
                continue

            fpath = os.path.join(full_dir, fname)
            iops, tail = parse_fio_log(fpath, wl)

            data_iops.setdefault(wl, {}).setdefault(bs, {cs: {nj: None for nj in NUMJOBS_LIST} for cs in CPU_SETS})
            data_tail.setdefault(wl, {}).setdefault(bs, {cs: {nj: None for nj in NUMJOBS_LIST} for cs in CPU_SETS})

            data_iops[wl][bs][cpu_set][njs] = iops
            data_tail[wl][bs][cpu_set][njs] = tail

    # 엑셀 생성
    wb = Workbook(); wb.remove(wb.active)

    # 시트 순서: workload 고정, bs는 BS_FILTERS 순서(있으면), 없으면 정렬
    for wl in WORKLOADS:
        bs_list = list(data_iops.get(wl, {}).keys())
        if not bs_list:
            continue
        if BS_FILTERS is not None:
            # 요청 순서 유지
            ordered = [b for b in BS_FILTERS if b in bs_list]
        else:
            ordered = sorted(bs_list, key=lambda x: (x[-1], int(x[:-1])) if x[:-1].isdigit() else x)
        for bs in ordered:
            write_workload_bs_sheet(
                wb,
                wl,
                bs,
                data_iops[wl][bs],
                data_tail[wl][bs]
            )

    wb.save(OUTPUT_XLSX)
    print(f"[DONE] wrote {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()

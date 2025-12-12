
import sys
from pathlib import Path
import re
from collections import defaultdict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side

TARGET_METRICS = ["IOPS", "BW(MB/s)", "lat_avg"]
WORKLOAD_ORDER = ["randread", "randwrite", "read", "write"]

def fs_name_from_path(p: Path) -> str:
    stem = p.stem  # e.g., "ext4_summary"
    m = re.match(r"(.+?)_summary$", stem, re.IGNORECASE)
    return m.group(1) if m else stem

def get_cpu_state(cpus: str):
    """Return 'BUSY' if cpus starts with '0', 'IDLE' if starts with '8', else None."""
    if isinstance(cpus, str):
        s = cpus.strip().strip("[]")
        if s.startswith("0"):
            return "BUSY"
        if s.startswith("8"):
            return "IDLE"
    return None

def parse_tables_from_sheet(ws):
    """
    Parse tables from the 'summary' sheet.
    Returns dict: {(cpus, workload, bs): {metric: [v1,v2,v3,v4]}}
    """
    max_row = ws.max_row
    cur = 1
    out = {}

    while cur <= max_row:
        head = ws.cell(cur, 2).value
        nj_ok = (
            cur + 1 <= max_row and
            ws.cell(cur+1, 2).value == 1 and
            ws.cell(cur+1, 3).value == 2 and
            ws.cell(cur+1, 4).value == 3 and
            ws.cell(cur+1, 5).value == 4
        )

        if isinstance(head, str) and nj_ok:
            # Normalize header and split
            head_norm = head.strip().strip("[]")
            parts = head_norm.split("_")
            if len(parts) >= 3:
                cpus, workload, bs = parts[0], parts[1], parts[2]
            else:
                cur += 1
                continue

            vals = {}
            for i, metric in enumerate(TARGET_METRICS, start=2):
                row_vals = [ws.cell(cur + i, c).value for c in range(2, 6)]
                vals[metric] = row_vals

            out[(cpus, workload, bs)] = vals

            # Skip header + numjobs row + 3 metric rows + 3 spacing rows
            cur += 1 + 1 + len(TARGET_METRICS) + 3
            continue
        else:
            cur += 1

    return out

def bs_key(bs: str):
    m = re.match(r"(\d+)\s*k$", bs.strip(), re.IGNORECASE)
    return int(m.group(1)) if m else 10**9

def workload_key(w):
    try:
        return WORKLOAD_ORDER.index(w)
    except ValueError:
        return len(WORKLOAD_ORDER) + (ord(w[0]) if w else 0)

def build_output(data_by_fs: dict, fs_order_pref=None):
    """
    data_by_fs: {fs_name: {(cpus, workload, bs): {metric: [v1..v4]}}}
    Produce tables keyed by (workload, bs), rows BUSY_<fs>, IDLE_<fs>.
    Returns:
      keys: sorted list of (workload, bs)
      table_map[(workload,bs,metric)][state][fs] = [v1..v4]
      fs_order: ordered list of fs names
    """
    wb_keys = set()
    for fs, d in data_by_fs.items():
        for (cpus, workload, bs) in d.keys():
            if get_cpu_state(cpus) in ("BUSY", "IDLE"):
                wb_keys.add((workload, bs))

    keys = sorted(wb_keys, key=lambda k: (workload_key(k[0]), bs_key(k[1])))

    table_map = defaultdict(lambda: defaultdict(dict))
    for fs, d in data_by_fs.items():
        for (workload, bs) in keys:
            for metric in TARGET_METRICS:
                for state in ("BUSY", "IDLE"):
                    table_map[(workload, bs, metric)].setdefault(state, {})
                    table_map[(workload, bs, metric)][state].setdefault(fs, [None]*4)

            for (cpus, w2, b2), metrics in d.items():
                if w2 != workload or b2 != bs:
                    continue
                state = get_cpu_state(cpus)
                if state not in ("BUSY", "IDLE"):
                    continue
                for metric in TARGET_METRICS:
                    table_map[(workload, bs, metric)][state][fs] = metrics.get(metric, [None]*4)

    all_fs = list(data_by_fs.keys())
    if fs_order_pref:
        fs_order = [fs for fs in fs_order_pref if fs in all_fs]
        fs_order += [fs for fs in sorted(all_fs) if fs not in fs_order]
    else:
        fs_order = sorted(all_fs)

    return keys, table_map, fs_order

def write_output(keys, table_map, fs_order, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    bold = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 1
    tables_written = 0

    for (workload, bs) in keys:
        for metric in TARGET_METRICS:
            header = f"{workload}_{bs}_{metric}"
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
            h = ws.cell(current_row, 2, header)
            h.font = bold; h.alignment = center
            current_row += 1

            for i, nj in enumerate([1,2,3,4], start=2):
                c = ws.cell(current_row, i, nj)
                c.font = bold; c.alignment = center
                c.border = border
            current_row += 1

            for state in ["BUSY", "IDLE"]:
                for fs in fs_order:
                    row_name = f"{state}_{fs}"
                    ws.cell(current_row, 1, row_name).font = bold
                    ws.cell(current_row, 1).alignment = right
                    ws.cell(current_row, 1).border = border

                    vals = table_map[(workload, bs, metric)][state].get(fs, [None]*4)
                    for i, v in enumerate(vals, start=2):
                        c = ws.cell(current_row, i, v)
                        c.alignment = center; c.border = border
                    current_row += 1

            current_row += 3
            tables_written += 1

    ws.column_dimensions["A"].width = 18
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 12

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[OK] Wrote {tables_written} tables -> {out_path}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 merge_fs_summaries_v3.py <fs1_summary.xlsx> [<fs2_summary.xlsx> ...] [--out OUTPUT.xlsx]")
        sys.exit(1)

    args = sys.argv[1:]
    if "--out" in args:
        idx = args.index("--out")
        out_path = Path(args[idx+1])
        files = [Path(p) for p in (args[:idx] + args[idx+2:])]
    else:
        files = [Path(p) for p in args]
        out_path = files[0].parent / "merged_summary_v3.xlsx"

    data_by_fs = {}
    for f in files:
        fs = fs_name_from_path(f)
        wb = load_workbook(f, data_only=True)
        if "summary" not in wb.sheetnames:
            print(f"[WARN] {f} has no 'summary' sheet. Skipped.")
            continue
        ws = wb["summary"]
        data_by_fs[fs] = parse_tables_from_sheet(ws)

    if not data_by_fs:
        print("[ERROR] No data parsed from inputs.")
        sys.exit(2)

    fs_pref = ["ext4", "fuse", "rfuse"]
    keys, table_map, fs_order = build_output(data_by_fs, fs_order_pref=fs_pref)
    write_output(keys, table_map, fs_order, out_path)

    # Sanity warnings for identical BUSY/IDLE rows
    for (workload, bs, metric), states in table_map.items():
        for fs in fs_order:
            b = states['BUSY'].get(fs)
            i = states['IDLE'].get(fs)
            if b and i and b == i and any(v is not None for v in (b + i)):
                print(f"[WARN] BUSY/IDLE identical for {fs} {workload}_{bs}_{metric}")

if __name__ == "__main__":
    main()


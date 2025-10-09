#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fio [workload]_[bs]_[numjobs].log → Excel 요약 스크립트
- 열(column): numjobs (상단 그룹 머리글은 [workload]_[bs]로 병합)
- 행(row): IOPS, BW, lat_avg, p95_lat, p99_lat, sys_cpu_usage, usr_cpu_usage
- 단일 시트에 모두 출력

Usage:
  python fio_to_xlsx.py --log-dir /path/to/logs --out /path/to/fio_summary.xlsx
Default:
  --log-dir . , --out ./fio_summary.xlsx
"""

import os
import re
import argparse
from typing import Dict, Tuple, Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

# ---------- Parsing helpers ----------

def parse_suffix_number(s: str) -> Optional[float]:
    """
    fio가 출력하는 수치(예: 10.2k, 3.1M 등)를 float로 변환 (k=1e3, M=1e6, G=1e9).
    쉼표 제거 및 대소문자 무시.
    """
    s = s.strip().lower().replace(',', '')
    m = re.match(r'^([0-9]*\.?[0-9]+)\s*([kmg])?$', s)
    if not m:
        try:
            return float(s)
        except Exception:
            return None
    val = float(m.group(1))
    suf = m.group(2)
    if suf == 'k':
        val *= 1_000
    elif suf == 'm':
        val *= 1_000_000
    elif suf == 'g':
        val *= 1_000_000_000
    return val

def to_us(value: float, unit: str) -> float:
    """
    시간 단위를 µs(마이크로초)로 통일.
    """
    unit = unit.strip().lower()
    if unit.startswith('n'):   # nsec
        return value / 1000.0
    if unit.startswith('u'):   # usec
        return value
    if unit.startswith('m'):   # msec
        return value * 1000.0
    if unit.startswith('s'):   # sec
        return value * 1_000_000.0
    return value  # 기본은 usec 가정

def extract_metrics_from_text(txt: str) -> Dict[str, float]:
    """
    fio 텍스트 로그에서 지표 추출:
      - IOPS
      - BW (MB/s로 통일; 괄호 안 (XX.XMB/s) 우선)
      - lat_avg (µs)
      - p95_lat, p99_lat (µs, clat percentiles 단위 기준)
      - usr/sys CPU (%)
    """
    metrics: Dict[str, float] = {}

    # IOPS
    m = re.search(r'IOPS\s*=\s*([0-9\.\,]+)\s*([kKmMgG])?', txt)
    if m:
        num = m.group(1)
        suf = (m.group(2) or '').lower()
        iops = parse_suffix_number(num + (suf if suf else ''))
        if iops is not None:
            metrics['IOPS'] = float(iops)

    # BW (MB/s) 우선, 없으면 MiB/s → MB/s 환산(1 MiB ≈ 1.048576 MB)
    m = re.search(r'BW\s*=\s*([0-9\.]+)\s*MiB/s\s*\(\s*([0-9\.]+)\s*MB/s\)', txt, re.IGNORECASE)
    if m:
        metrics['BW'] = float(m.group(2))
    else:
        m = re.search(r'BW\s*=\s*([0-9\.]+)\s*MiB/s', txt, re.IGNORECASE)
        if m:
            mib = float(m.group(1))
            metrics['BW'] = mib * 1.048576
        else:
            m = re.search(r'BW\s*=\s*([0-9\.]+)\s*MB/s', txt, re.IGNORECASE)
            if m:
                metrics['BW'] = float(m.group(1))

    # 평균 지연(lat_avg): lat(...) 또는 clat(...) 블록의 avg
    m = re.search(r'\n\s*lat\s*\((nsec|usec|msec|sec)\)\s*:\s*.*?avg\s*=\s*([0-9\.]+)', txt,
                  re.IGNORECASE | re.DOTALL)
    if m:
        unit = m.group(1)
        avg = float(m.group(2))
        metrics['lat_avg'] = to_us(avg, unit)
    else:
        m = re.search(r'\n\s*clat\s*\((nsec|usec|msec|sec)\)\s*:\s*.*?avg\s*=\s*([0-9\.]+)', txt,
                      re.IGNORECASE | re.DOTALL)
        if m:
            unit = m.group(1)
            avg = float(m.group(2))
            metrics['lat_avg'] = to_us(avg, unit)

    # clat percentiles 단위 파악 (p95/p99 추출용)
    punit = None
    m = re.search(r'clat percentiles\s*\(\s*(nsec|usec|msec|sec)\s*\)', txt, re.IGNORECASE)
    if m:
        punit = m.group(1).lower()

    # 95th, 99th
    m = re.search(r'95\.00th=\[\s*([0-9\.]+)\s*\]', txt)
    if m and punit:
        metrics['p95_lat'] = to_us(float(m.group(1)), punit)

    m = re.search(r'99\.00th=\[\s*([0-9\.]+)\s*\]', txt)
    if m and punit:
        metrics['p99_lat'] = to_us(float(m.group(1)), punit)

    # CPU
    m = re.search(r'cpu\s*:\s*usr=([0-9\.]+)%,\s*sys=([0-9\.]+)%', txt, re.IGNORECASE)
    if m:
        metrics['usr_cpu_usage'] = float(m.group(1))
        metrics['sys_cpu_usage'] = float(m.group(2))

    return metrics

# ---------- Excel builder ----------

def write_excel(data: Dict[Tuple[str, str], Dict[int, Dict[str, float]]], out_path: str) -> None:
    """
    data 구조:
      {
        (workload, bs): {
           numjobs(int): {
             'IOPS': float, 'BW': float, 'lat_avg': float,
             'p95_lat': float, 'p99_lat': float,
             'sys_cpu_usage': float, 'usr_cpu_usage': float
           }, ...
        }, ...
      }
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "fio_summary"

    metric_rows = ["IOPS", "BW", "lat_avg", "p95_lat", "p99_lat", "sys_cpu_usage", "usr_cpu_usage"]
    row_titles = {
        "IOPS": "IOPS",
        "BW": "BW (MB/s)",
        "lat_avg": "lat_avg (µs)",
        "p95_lat": "p95_lat (µs)",
        "p99_lat": "p99_lat (µs)",
        "sys_cpu_usage": "sys_cpu_usage (%)",
        "usr_cpu_usage": "usr_cpu_usage (%)",
    }

    # 헤더: A1:A2 = "metric" 병합, 1행=그룹헤더(merged), 2행=numjobs
    ws.cell(row=1, column=1, value="metric")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True)

    # 그룹 헤더 및 numjobs 컬럼 생성
    col = 2
    for (workload, bs) in sorted(data.keys()):
        group = data[(workload, bs)]
        group_name = f"{workload}_{bs}"
        nums = sorted(group.keys())
        ws.cell(row=1, column=col, value=group_name)
        if len(nums) > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(nums) - 1)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=col).font = Font(bold=True)

        for nj in nums:
            ws.cell(row=2, column=col, value=nj)
            ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=2, column=col).font = Font(bold=True)
            col += 1

    # 데이터 영역
    start_data_row = 3
    for i, met in enumerate(metric_rows):
        r = start_data_row + i
        ws.cell(row=r, column=1, value=row_titles[met]).font = Font(bold=True)
        c = 2
        for (workload, bs) in sorted(data.keys()):
            nums = sorted(data[(workload, bs)].keys())
            for nj in nums:
                val = data[(workload, bs)][nj].get(met, None)
                ws.cell(row=r, column=c, value=val)
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
                c += 1

    # 스타일: 테두리, 컬럼 너비, 프리즈
    ws.freeze_panes = "B3"
    thin = Side(border_style="thin", color="000000")
    for r in ws.iter_rows(min_row=1, max_row=start_data_row + len(metric_rows) - 1,
                          min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for c in range(1, ws.max_column + 1):
        maxlen = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(10, maxlen + 2), 40)

    wb.save(out_path)

# ---------- Main ----------

def main():
    ap = argparse.ArgumentParser(description="Parse fio logs to a single-sheet Excel summary.")
    ap.add_argument("--log-dir", default=".", help="Directory containing [workload]_[bs]_[numjobs].log files")
    ap.add_argument("--out", default="./fio_summary.xlsx", help="Output Excel path")
    args = ap.parse_args()

    pat = re.compile(r'^(?P<workload>[^_]+)_(?P<bs>[^_]+)_(?P<numjobs>\d+)\.log$')

    data: Dict[Tuple[str, str], Dict[int, Dict[str, float]]] = {}
    for fname in os.listdir(args.log_dir):
        if not fname.endswith(".log"):
            continue
        m = pat.match(fname)
        if not m:
            # 패턴 매칭 안 되면 스킵
            continue
        workload = m.group('workload')
        bs = m.group('bs')
        numjobs = int(m.group('numjobs'))
        key = (workload, bs)
        fpath = os.path.join(args.log_dir, fname)
        with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
            txt = f.read()
        metrics = extract_metrics_from_text(txt)
        data.setdefault(key, {})[numjobs] = metrics

    if not data:
        raise SystemExit("No matching fio logs found. Expect files like [workload]_[bs]_[numjobs].log")

    write_excel(data, args.out)
    print(f"Written: {args.out}")

if __name__ == "__main__":
    main()


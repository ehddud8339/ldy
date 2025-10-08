#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
merge_io_xlsx.py

세 개의 결과 엑셀(ext4/fuse/rfuse)을 읽어서
(section, bs)별 시트에 numjobs를 열로, 파일시스템별 핵심 지표를 행으로 정리합니다.

- 입력: ext4_results.xlsx, fuse_results.xlsx, rfuse_results.xlsx
  (앞서 만든 "블록 쌓기" 포맷: 병합제목 [section]_[bs] → 헤더(열=numjobs) → 지표행)
- 출력: 섹션×블록크기별 시트 (예: seqread-4k, randwrite-128k)
  행: ext4/fuse/rfuse의 IOPS/BW/lat_avg/p95/p99 (각 FS 블록 사이에 빈 줄)
  열: numjobs (세 엑셀의 numjobs 합집합)
- read 계열(seqread/randread/read)은 read 전용 지표를, write 계열은 write 전용 지표를 사용
- NaN/결측치는 빈 칸 처리
"""

import argparse
import math
from pathlib import Path
from typing import Dict, Tuple, List, Any, Optional

import pandas as pd
from openpyxl import load_workbook

# ----------------------------
# Helpers: parse block-stacked XLSX (our previous format)
# ----------------------------

def parse_blocks_from_sheet(xlsx_path: Path) -> Dict[Tuple[str, str], pd.DataFrame]:
    """
    Read first (only) sheet of our stacked-block workbook and return:
    {
      (section, bs): DataFrame,   # rows = metric names, cols = numjobs(int), values=float|NaN
      ...
    }
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]  # 첫 시트 사용

    blocks: Dict[Tuple[str, str], pd.DataFrame] = {}

    row_idx = 1  # 1-based in openpyxl
    maxcol = ws.max_column
    maxrow = ws.max_row

    while row_idx <= maxrow:
        title_cell = ws.cell(row=row_idx, column=1).value
        if not title_cell:
            row_idx += 1
            continue

        # 제목: "seqread_4k" 형태
        title = str(title_cell).strip()
        if "_" not in title:
            row_idx += 1
            continue
        section, bs = title.split("_", 1)

        # 헤더(바로 다음 행): numjobs
        header_row = row_idx + 1
        numjobs: List[int] = []
        col = 2
        while True:
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                break
            try:
                numjobs.append(int(val))
            except Exception:
                break
            col += 1
        if not numjobs:
            row_idx += 1
            continue

        # 데이터 행들
        data_start = row_idx + 2
        r = data_start
        metrics: List[str] = []
        matrix: Dict[int, List[Optional[float]]] = {nj: [] for nj in numjobs}

        while r <= maxrow:
            metric_name = ws.cell(row=r, column=1).value
            if metric_name is None or str(metric_name).strip() == "":
                break
            metric_name = str(metric_name).strip()

            metrics.append(metric_name)
            for c_idx, nj in enumerate(numjobs, start=2):
                v = ws.cell(row=r, column=c_idx).value
                if isinstance(v, (int, float)):
                    if math.isfinite(float(v)):
                        matrix[nj].append(float(v))
                    else:
                        matrix[nj].append(float("nan"))
                else:
                    matrix[nj].append(float("nan"))
            r += 1

        if metrics:
            df = pd.DataFrame({nj: matrix[nj] for nj in numjobs}, index=metrics)
            blocks[(section, bs)] = df

        # 다음 블록으로 (블록 종료 후 한 줄 비움)
        row_idx = r + 1

    return blocks


def load_fs_book(path: Path) -> Dict[Tuple[str, str], pd.DataFrame]:
    """Load one FS result workbook (ext4/fuse/rfuse) and return blocks dict."""
    return parse_blocks_from_sheet(path)


# ----------------------------
# Merge triplet (ext4/fuse/rfuse) into per-(section,bs) sheets
# ----------------------------

READ_SECTIONS = ("seqread", "randread", "read")
WRITE_SECTIONS = ("seqwrite", "randwrite", "write")

def is_read_section(section: str) -> bool:
    s = section.lower()
    return any(s.startswith(x) for x in READ_SECTIONS)

def is_write_section(section: str) -> bool:
    s = section.lower()
    return any(s.startswith(x) for x in WRITE_SECTIONS)

def pick_metrics_for_section(section: str) -> Tuple[str, str, str]:
    """
    Return metric field names for (lat_avg, p95, p99) depending on read/write section.
    """
    if is_read_section(section):
        return ("avg_read_us", "p95_read_us", "p99_read_us")
    else:
        return ("avg_write_us", "p95_write_us", "p99_write_us")

def extract_triplet_for_block(section: str, df: pd.DataFrame) -> Dict[str, pd.Series]:
    """
    From a block DF (rows=metric names, cols=numjobs), return:
    {
      "IOPS": Series(numjobs),
      "BW_MBps": Series(numjobs),
      "lat_avg_us": Series(numjobs),
      "p95_us": Series(numjobs),
      "p99_us": Series(numjobs)
    }
    """
    lat_avg_name, p95_name, p99_name = pick_metrics_for_section(section)

    def get_row(name: str) -> pd.Series:
        if name in df.index:
            return pd.to_numeric(df.loc[name], errors="coerce")
        return pd.Series([math.nan] * df.shape[1], index=df.columns)

    out = {
        "IOPS": get_row("IOPS_total"),
        "BW_MBps": get_row("BW_total_MBps"),
        "lat_avg_us": get_row(lat_avg_name),
        "p95_us": get_row(p95_name),
        "p99_us": get_row(p99_name),
    }
    return out

def build_sheet_matrix(section: str,
                       ext4_blk: Optional[pd.DataFrame],
                       fuse_blk: Optional[pd.DataFrame],
                       rfuse_blk: Optional[pd.DataFrame]) -> Tuple[List[int], pd.DataFrame]:
    """
    Build final matrix rows for one (section, bs) across 3 FS.
    Rows order:
      ext4_IOPS, ext4_BW_MBps, ext4_lat_avg_us, ext4_p95_us, ext4_p99_us,
      (blank),
      fuse_IOPS, ...
      (blank),
      rfuse_IOPS, ...
    Columns: union of numjobs across available blocks.
    """
    fs_map = {}
    if ext4_blk is not None:
        fs_map["ext4"] = extract_triplet_for_block(section, ext4_blk)
    if fuse_blk is not None:
        fs_map["fuse"] = extract_triplet_for_block(section, fuse_blk)
    if rfuse_blk is not None:
        fs_map["rfuse"] = extract_triplet_for_block(section, rfuse_blk)

    if not fs_map:
        return [], pd.DataFrame()

    # union of numjobs
    union_cols = set()
    for g in fs_map.values():
        for s in g.values():
            union_cols.update(map(int, s.index.tolist()))
    union_cols = sorted(union_cols)

    # build rows
    row_labels: List[str] = []
    rows: List[List[Optional[float]]] = []

    def append_fs_block(fsname: str, group: Dict[str, pd.Series]):
        mapping_order = [("IOPS", "IOPS"),
                         ("BW_MBps", "BW_MBps"),
                         ("lat_avg_us", "lat_avg_us"),
                         ("p95_us", "p95_us"),
                         ("p99_us", "p99_us")]
        for key, label in mapping_order:
            row_labels.append(f"{fsname}_{label}")
            s = group[key].copy()
            s.index = s.index.astype(int)
            s = s.reindex(union_cols)
            rows.append([ (float(v) if isinstance(v, (int,float)) and math.isfinite(v) else None) for v in s.tolist() ])
        # blank separator
        row_labels.append("")
        rows.append([None] * len(union_cols))

    for fsname in ["ext4", "fuse", "rfuse"]:
        if fsname in fs_map:
            append_fs_block(fsname, fs_map[fsname])

    # drop trailing blank if last block appended one
    if row_labels and row_labels[-1] == "":
        row_labels.pop()
        rows.pop()

    mat = pd.DataFrame(rows, index=row_labels, columns=union_cols)
    return union_cols, mat


# ----------------------------
# Main
# ----------------------------

def main():
    ap = argparse.ArgumentParser(description="Merge ext4/fuse/rfuse XLSX into consolidated per-(section,bs) sheets.")
    ap.add_argument("--ext4", required=True, help="Path to ext4_results.xlsx")
    ap.add_argument("--fuse", required=True, help="Path to fuse_results.xlsx")
    ap.add_argument("--rfuse", required=True, help="Path to rfuse_results.xlsx")
    ap.add_argument("-o", "--output", default="merged_results.xlsx", help="Output XLSX")
    args = ap.parse_args()

    ext4_blocks = load_fs_book(Path(args.ext4))
    fuse_blocks = load_fs_book(Path(args.fuse))
    rfuse_blocks = load_fs_book(Path(args.rfuse))

    # gather all (section, bs) keys
    all_keys = set(ext4_blocks.keys()) | set(fuse_blocks.keys()) | set(rfuse_blocks.keys())
    if not all_keys:
        print("No blocks found in input workbooks.")
        return

    out_path = Path(args.output).resolve()
    with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
        wb = writer.book
        for (section, bs) in sorted(all_keys):
            sheet_name = f"{section}-{bs}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            ext4_blk = ext4_blocks.get((section, bs))
            fuse_blk = fuse_blocks.get((section, bs))
            rfuse_blk = rfuse_blocks.get((section, bs))

            cols, mat = build_sheet_matrix(section, ext4_blk, fuse_blk, rfuse_blk)
            if mat.empty:
                continue

            # 새 워크시트
            ws = wb.add_worksheet(sheet_name)
            writer.sheets[sheet_name] = ws

            header_fmt = wb.add_format({"bold": True, "align": "center", "valign": "vcenter", "border": 1})
            idx_fmt    = wb.add_format({"bold": True, "align": "left",   "valign": "vcenter", "border": 1})
            num_fmt    = wb.add_format({"border": 1, "num_format": "0.00"})
            int_fmt    = wb.add_format({"border": 1, "num_format": "0"})
            cell_fmt   = wb.add_format({"border": 1})

            ws.set_column(0, 0, 22)              # 라벨
            ws.set_column(1, len(cols), 12)      # 수치

            # 헤더 행: "", numjobs...
            ws.write(0, 0, "", header_fmt)
            for i, nj in enumerate(cols, start=1):
                ws.write(0, i, int(nj), header_fmt)

            # 본문: ★ 수정안 A — 빈 라벨/Series 안전 처리
            for r, label in enumerate(mat.index.tolist(), start=1):
                # 좌측 라벨
                ws.write(r, 0, label, idx_fmt if label else cell_fmt)
                for c, nj in enumerate(cols, start=1):
                    if not label:
                        # 구분용 빈 줄: 전열 빈칸
                        ws.write_blank(r, c, None, cell_fmt)
                        continue
                    v = mat.loc[label, nj]
                    # 중복 인덱스로 Series가 반환되면 첫 값만 사용
                    if hasattr(v, "iloc"):
                        v = v.iloc[0] if len(v) > 0 else None
                    # NaN/None → 빈칸
                    if v is None or (isinstance(v, float) and not math.isfinite(v)):
                        ws.write_blank(r, c, None, cell_fmt)
                    else:
                        fmt = int_fmt if label.endswith("_IOPS") else num_fmt
                        ws.write(r, c, float(v), fmt)

    print(f"[ok] wrote: {out_path}")

if __name__ == "__main__":
    main()


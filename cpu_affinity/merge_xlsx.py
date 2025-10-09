#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
세 가지 xlsx(ext4/fuse/rfuse) → summary.xlsx
- 입력 각 파일은 'fio_summary' 시트, 상단 구조:
  Row1: [workload]_[bs] (numjobs 수 만큼 병합)
  Row2: numjobs 값들
  Row>=3: 지표 행 (예: IOPS, BW (MB/s), lat_avg (µs), ...)
- 출력 summary.xlsx:
  * 시트: [workload]
  * 열(2중 헤더): 1행 병합 헤더 = [bs]_[row], 2행 = numjobs
  * 행: ext4, fuse, rfuse

Usage:
  python merge_fio_results.py \
      --ext4 /path/to/ext4_results.xlsx \
      --fuse /path/to/fuse_results.xlsx \
      --rfuse /path/to/rfuse_results.xlsx \
      --out  /path/to/summary.xlsx
"""

import argparse
from collections import defaultdict, OrderedDict
from typing import Dict, Tuple, List, Any, Set
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side

# --------- 유틸: 지표명 정규화 ---------
def norm_metric(label: str) -> str:
    """입력 시트의 행 제목을 표준 토큰으로 정규화."""
    s = (label or "").strip().lower()
    # 우선순위: 키워드 포함 여부로 매핑
    if s.startswith("iops"):
        return "IOPS"
    if s.startswith("bw"):
        return "BW"
    if s.startswith("lat_avg"):
        return "lat_avg"
    if s.startswith("p95"):
        return "p95_lat"
    if s.startswith("p99"):
        return "p99_lat"
    if s.startswith("sys_cpu"):
        return "sys_cpu_usage"
    if s.startswith("usr_cpu"):
        return "usr_cpu_usage"
    # 여유 케이스: 괄호/단위 포함 라벨들
    if "lat_avg" in s:
        return "lat_avg"
    if "p95" in s:
        return "p95_lat"
    if "p99" in s:
        return "p99_lat"
    if "sys" in s and "cpu" in s:
        return "sys_cpu_usage"
    if "usr" in s and "cpu" in s:
        return "usr_cpu_usage"
    return label  # 미정규화 라벨은 그대로

# 출력 지표 순서
METRIC_ORDER = ["IOPS", "BW", "lat_avg", "p95_lat", "p99_lat", "sys_cpu_usage", "usr_cpu_usage"]

# --------- 단일 입력 파일(fio_summary) 파서 ---------
def parse_fio_summary_sheet(path: str, sheet_name: str = "fio_summary"):
    """
    입력 파일에서 아래 구조로 반환:
    data[workload][bs][metric][numjobs] = value
    또, 각 workload에 대해 사용된 numjobs 집합과 bs 집합도 함께 반환.
    """
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"{path}: sheet '{sheet_name}' not found")
    ws = wb[sheet_name]

    # col->(group, numjobs) 매핑: group = "[workload]_[bs]"
    col_map: Dict[int, Tuple[str, Any]] = {}
    last_group = None
    for col in range(2, ws.max_column + 1):
        g = ws.cell(row=1, column=col).value
        if g is not None:
            last_group = str(g)
        numj = ws.cell(row=2, column=col).value
        col_map[col] = (last_group, numj)

    # 데이터 적재
    data: Dict[str, Dict[str, Dict[str, Dict[Any, Any]]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    workload_set: Set[str] = set()
    bs_set_by_workload: Dict[str, Set[str]] = defaultdict(set)
    nj_set_by_workload_bs: Dict[Tuple[str, str], Set[Any]] = defaultdict(set)

    for row in range(3, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        metric = norm_metric(str(label) if label is not None else "")
        for col in range(2, ws.max_column + 1):
            group, numj = col_map[col]
            if not group:
                continue
            # group = "[workload]_[bs]": workload와 bs 분리
            parts = str(group).split("_", 1)
            if len(parts) != 2:
                # 포맷이 다르면 스킵
                continue
            workload, bs = parts[0], parts[1]
            val = ws.cell(row=row, column=col).value
            if metric:
                data[workload][bs][metric][numj] = val
                workload_set.add(workload)
                bs_set_by_workload[workload].add(bs)
                nj_set_by_workload_bs[(workload, bs)].add(numj)

    return data, workload_set, bs_set_by_workload, nj_set_by_workload_bs

# --------- 세 파일 결합 ---------
def merge_three_sources(ext4_path: str, fuse_path: str, rfuse_path: str):
    """
    세 파일을 파싱하여 통합 데이터 구성:
      merged[workload][bs][metric][numjobs][fs] = value
    또한 각 workload별 사용된 bs, (workload, bs)별 numjobs 집합도 정리.
    """
    sources = {
        "ext4": ext4_path,
        "fuse": fuse_path,
        "rfuse": rfuse_path,
    }

    merged: Dict[str, Dict[str, Dict[str, Dict[Any, Dict[str, Any]]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    )
    bs_by_workload: Dict[str, Set[str]] = defaultdict(set)
    nj_by_workload_bs: Dict[Tuple[str, str], Set[Any]] = defaultdict(set)
    workloads_all: Set[str] = set()

    for fs, path in sources.items():
        data, w_set, bs_set_by_w, nj_set_by_wb = parse_fio_summary_sheet(path)
        workloads_all |= w_set
        for w in w_set:
            bs_by_workload[w] |= bs_set_by_w[w]
        for (w, b), njs in nj_set_by_wb.items():
            nj_by_workload_bs[(w, b)] |= njs

        # 병합
        for w, bsd in data.items():
            for b, md in bsd.items():
                for m, njd in md.items():
                    for nj, val in njd.items():
                        merged[w][b][m][nj][fs] = val

    return merged, workloads_all, bs_by_workload, nj_by_workload_bs

# --------- summary.xlsx 작성 ---------
def write_summary_xlsx(
    merged, workloads: Set[str], bs_by_workload: Dict[str, Set[str]],
    nj_by_workload_bs: Dict[Tuple[str, str], Set[Any]],
    out_path: str
):
    """
    시트 = workload
    1행: 병합 헤더 [bs]_[metric]
    2행: numjobs
    3행~: ext4, fuse, rfuse
    """
    wb = Workbook()
    # 기본 생성 시트는 첫 workload로 교체, 없으면 그대로
    created_first = False

    FS_ORDER = ["ext4", "fuse", "rfuse"]

    for w in sorted(workloads):
        ws = wb.active if not created_first else wb.create_sheet(title=w)
        if not created_first:
            ws.title = w
            created_first = True

        # 헤더 쓰기
        ws.cell(row=1, column=1, value="filesystem")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(bold=True)

        col = 2
        # bs 순서 고정(알파/숫자 섞임 가능 → 문자열 정렬)
        for bs in sorted(bs_by_workload.get(w, [])):
            # numjobs 집합(모든 FS 합집합)
            nj_list = sorted(nj_by_workload_bs.get((w, bs), []), key=lambda x: (isinstance(x, str), x))
            # 각 metric별 그룹을 순회
            for metric in METRIC_ORDER:
                header = f"{bs}_{metric}"
                # 병합 헤더
                if len(nj_list) > 1:
                    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(nj_list) - 1)
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=1, column=col).font = Font(bold=True)
                # 2행: numjobs
                for nj in nj_list:
                    ws.cell(row=2, column=col, value=nj)
                    ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    ws.cell(row=2, column=col).font = Font(bold=True)
                    col += 1

        # 데이터(행 = filesystem)
        row_start = 3
        for i, fs in enumerate(FS_ORDER):
            r = row_start + i
            ws.cell(row=r, column=1, value=fs).font = Font(bold=True)
            c = 2
            for bs in sorted(bs_by_workload.get(w, [])):
                nj_list = sorted(nj_by_workload_bs.get((w, bs), []), key=lambda x: (isinstance(x, str), x))
                for metric in METRIC_ORDER:
                    # 이 (w, bs, metric)에 대해 numjobs 순서대로 값 채움
                    for nj in nj_list:
                        val = merged.get(w, {}).get(bs, {}).get(metric, {}).get(nj, {}).get(fs, None)
                        ws.cell(row=r, column=c, value=val)
                        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
                        c += 1

        # 스타일(테두리/폭/프리즈)
        thin = Side(border_style="thin", color="000000")
        max_row = row_start + len(FS_ORDER) - 1
        max_col = ws.max_column
        for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in r:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # 보기 편의: 컬럼 폭
        for c in range(1, max_col + 1):
            maxlen = 0
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                maxlen = max(maxlen, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = min(max(10, maxlen + 2), 40)

        ws.freeze_panes = "B3"

    wb.save(out_path)
    print(f"Written: {out_path}")

# --------- main ---------
def main():
    ap = argparse.ArgumentParser(description="Merge ext4/fuse/rfuse fio Excel into summary.xlsx (by workload).")
    ap.add_argument("--ext4", required=True, help="ext4_results.xlsx path")
    ap.add_argument("--fuse", required=True, help="fuse_results.xlsx path")
    ap.add_argument("--rfuse", required=True, help="rfuse_results.xlsx path")
    ap.add_argument("--out",  required=True, help="output summary.xlsx path")
    args = ap.parse_args()

    merged, workloads, bs_by_workload, nj_by_workload_bs = merge_three_sources(
        args.ext4, args.fuse, args.rfuse
    )
    if not workloads:
        raise SystemExit("No workloads found in input files.")
    write_summary_xlsx(merged, workloads, bs_by_workload, nj_by_workload_bs, args.out)

if __name__ == "__main__":
    main()


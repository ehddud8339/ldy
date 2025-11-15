#!/usr/bin/env python3
import os
import glob
import argparse
from openpyxl import Workbook

# 관심 있는 지표 목록
METRICS = [
    "pgfault", "pgmajfault",
    #"nr_dirty", "nr_writeback",
    "pgpgin", "pgpgout",
    #"numa_hit", "numa_miss", "numa_local", "numa_other", "numa_interleave",
    "nr_dirtied", "nr_written",
    "pswpin", "pswpout",
    #"allocstall_dma", "allocstall_dma32", "allocstall_normal", "allocstall_movable",
]

PAGE_SIZE = 1024  # /proc/vmstat의 pgpgin/pgpgout 단위(보통 1KB 블록)


def read_vmstat_file(path):
    """
    /proc/vmstat 형식 로그 파일을 {key: int(value)} dict로 파싱
    """
    data = {}
    with open(path, "r") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) != 2:
                continue
            key, val = parts
            try:
                data[key] = int(val)
            except ValueError:
                continue
    return data


def diff_vmstat(before, after):
    """
    before/after dict를 받아 after - before diff dict로 변환
    """
    diff = {}
    keys = set(before.keys()) | set(after.keys())
    for k in keys:
        b = before.get(k, 0)
        a = after.get(k, 0)
        diff[k] = a - b
    return diff


def collect_rows(input_dir):
    """
    input_dir 내의 *_vmstat_before/after.log를 찾아
    bound -> {sched -> diff_dict} 구조로 리턴.

    파일 이름 예:
      rr_io_bound_vmstat_before.log
      rr_io_bound_vmstat_after.log

    => bound = "io_bound", sched = "rr"
    """
    # rows[bound][sched] = diff_dict
    rows = {}

    before_files = glob.glob(os.path.join(input_dir, "*_vmstat_before.log"))

    for before_path in sorted(before_files):
        base_name = os.path.basename(before_path)
        base_prefix = base_name.replace("_vmstat_before.log", "")
        after_path = os.path.join(input_dir, base_prefix + "_vmstat_after.log")

        if not os.path.exists(after_path):
            print(f"[SKIP] No matching after log for: {before_path}")
            continue

        # base_prefix: "rr_mem_bound" 같은 형태라고 가정
        if "_" not in base_prefix:
            print(f"[WARN] Unexpected prefix (no underscore): {base_prefix}")
            sched = base_prefix
            bound = "default"
        else:
            sched, bound = base_prefix.split("_", 1)

        before = read_vmstat_file(before_path)
        after = read_vmstat_file(after_path)
        diff = diff_vmstat(before, after)

        if bound not in rows:
            rows[bound] = {}
        rows[bound][sched] = diff

    return rows


def write_xlsx(bound_rows, output_path):
    """
    bound_rows: {bound: {sched: diff_dict}}
    각 bound마다 다른 sheet를 만들고,
    sheet 안에서는 행: sched, 열: METRICS (+ *_bytes) 구조로 만든다.
    """
    wb = Workbook()
    first_sheet = True

    # 각 bound에 대해 sheet 하나씩
    for bound in sorted(bound_rows.keys()):
        sched_map = bound_rows[bound]

        if not sched_map:
            continue

        if first_sheet:
            ws = wb.active
            ws.title = bound
            first_sheet = False
        else:
            ws = wb.create_sheet(title=bound)

        # 헤더 구성: 첫 컬럼은 sched 이름
        headers = ["name"]  # 여기 name에는 sched만 들어감
        headers.extend(METRICS)

        # pgpgin/pgpgout는 byte 버전 컬럼도 추가
        extra_metrics = []
        for m in METRICS:
            if m in ("pgpgin", "pgpgout"):
                extra_metrics.append(m + "_bytes")
        headers.extend(extra_metrics)

        # 헤더 쓰기
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)

        # 데이터 행 작성
        row_idx = 2
        for sched in sorted(sched_map.keys()):
            diff = sched_map[sched]
            col_idx = 1

            # name 컬럼 = sched
            ws.cell(row=row_idx, column=col_idx, value=sched)
            col_idx += 1

            metric_values = {}
            # 기본 메트릭 값들
            for m in METRICS:
                val = diff.get(m, 0)
                metric_values[m] = val
                ws.cell(row=row_idx, column=col_idx, value=val)
                col_idx += 1

            # pgpgin/pgpgout bytes 컬럼
            for m in METRICS:
                if m in ("pgpgin", "pgpgout"):
                    val = metric_values.get(m, 0)
                    ws.cell(row=row_idx, column=col_idx, value=val * PAGE_SIZE)
                    col_idx += 1

            row_idx += 1

    wb.save(output_path)
    print(f"[OK] Wrote XLSX: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate XLSX summary from vmstat before/after logs."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Directory containing *_vmstat_before.log and *_vmstat_after.log",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Output XLSX file path (e.g., ./vmstat_summary.xlsx)",
    )

    args = parser.parse_args()

    input_dir = os.path.abspath(args.input)
    output_path = os.path.abspath(args.output)

    if not os.path.isdir(input_dir):
        print(f"[ERROR] Input directory does not exist or is not a directory: {input_dir}")
        return

    bound_rows = collect_rows(input_dir)
    if not bound_rows:
        print(f"[ERROR] No before/after vmstat pairs found in {input_dir}")
        return

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    write_xlsx(bound_rows, output_path)


if __name__ == "__main__":
    main()
